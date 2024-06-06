# Load the AdSec API
import os
import tempfile
from random import randint

# importing packages
import openpyxl as op  # OPen source library that can be used for reading and writing into the excel files
from Oasys.AdSec import IAdSec, ILoad, ISection
from Oasys.AdSec.DesignCode import EN1992
from Oasys.AdSec.StandardMaterials import Steel
from Oasys.Profiles import ICircleProfile
from OasysUnits import Force, Length, Moment
from OasysUnits.Units import ForceUnit, LengthUnit, MomentUnit


def main():
    # Part1
    # Getting the workbook from the excel file.
    excelfile_path = create_excel_with_random_loads()

    # Part2
    # Open a workbook with worksheet with name "Sheet1"
    workbook = op.load_workbook(excelfile_path)
    worksheet = workbook["Sheet1"]

    # creating adsec object and defining section material
    section_material = Steel.EN1993.Edition_2005.S235
    ad_sec = IAdSec.Create(EN1992.Part1_1.Edition_2004.NationalAnnex.GB.Edition_2014)

    section = create_section(section_material)

    # Analysing the section
    circle_section_analysis = ad_sec.Analyse(section)

    max_utilisation = 100

    for i in range(1, 11):
        circle_load = create_circle_load(worksheet, i)

        # Getting the strength results
        strength_result = circle_section_analysis.Strength.Check(circle_load)
        utilisation = round(strength_result.LoadUtilisation.Percent, 1)
        uls_status = "Clear" if utilisation < max_utilisation else "Not Clear"
        worksheet[f"D{i}"].value = utilisation
        worksheet[f"E{i}"].value = uls_status

    # excel file will be created in the temp folder
    workbook.save(excelfile_path)


def create_excel_with_random_loads():
    temp_fd, temp_path = tempfile.mkstemp(prefix="OUTPUT", suffix=".xlsx")
    os.close(temp_fd)
    workbook = op.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    # adding random values into cells
    for i in range(1, 11):
        worksheet.cell(i, 1, value=randint(-100, 100) * 1000)
        worksheet.cell(i, 2, value=randint(-100, 100) * 1000)
        worksheet.cell(i, 3, value=randint(-100, 100) * 1000)

    workbook.save(temp_path)

    return temp_path


def create_circle_load(worksheet, i):
    # Getting the loads and converting it
    force_x = worksheet[f"A{i}"].value
    moment_yy = worksheet[f"B{i}"].value
    moment_zz = worksheet[f"C{i}"].value
    circle_load = ILoad.Create(
        Force(force_x, ForceUnit.Kilonewton),
        Moment(moment_yy, MomentUnit.KilonewtonMeter),
        Moment(moment_zz, MomentUnit.KilonewtonMeter),
    )
    return circle_load


def create_section(section_material):
    diameter = Length(80, LengthUnit.Inch)
    circle_profile = ICircleProfile.Create(diameter)
    section = ISection.Create(circle_profile, section_material)
    return section


if __name__ == "__main__":
    main()
